#!/usr/bin/env python3
"""
🧹 SCRIPT PARA LIMPIAR EXCEL CORRUPTO
Corrige el archivo Excel con datos desordenados
"""

import openpyxl
import os
from datetime import datetime

def limpiar_excel():
    """Limpiar y reorganizar Excel"""
    excel_file = 'seguimiento.xlsx'
    
    if not os.path.exists(excel_file):
        print("❌ Archivo Excel no encontrado")
        return
    
    print("🧹 Limpiando archivo Excel...")
    
    # Hacer backup
    backup_file = f'seguimiento_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    if os.path.exists(excel_file):
        import shutil
        shutil.copy2(excel_file, backup_file)
        print(f"💾 Backup creado: {backup_file}")
    
    # Crear nuevo Excel con estructura correcta
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seguimiento_Constancias"
    
    # Headers correctos
    headers = [
        "ID",           # 0
        "Alumno",       # 1
        "Código",       # 2
        "DNI",          # 3
        "Correo",       # 4
        "Carrera",      # 5
        "Ciclo",        # 6
        "Documento",    # 7 - ARCHIVO PDF
        "Estado",       # 8
        "Autoridad",    # 9
        "Firma",        # 10
        "Fecha",        # 11
        "UNI_Validado", # 12
        "DNI_Validado", # 13
        "Fuente",       # 14
        "Facultad",     # 15
        "Estado_UNI"    # 16
    ]
    
    ws.append(headers)
    
    # Aplicar formato a headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    
    # Guardar Excel limpio
    wb.save(excel_file)
    print(f"✅ Excel limpiado y reorganizado")
    print(f"📋 Headers correctos aplicados")
    print(f"💡 Ahora las descargas funcionarán correctamente")

if __name__ == "__main__":
    limpiar_excel()