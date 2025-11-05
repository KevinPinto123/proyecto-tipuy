from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import os
import uuid
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import time

# Importar servicios de validación
from uni_validation_service import UNIValidationService
from dni_validation_service import DNIValidationService

class RPAService:
    def __init__(self):
        self.excel_file = "seguimiento.xlsx"
        self.pdf_folder = "autoridad_entrada"
        self.uni_validator = UNIValidationService()
        self.dni_validator = DNIValidationService()
        self._inicializar_excel()
    
    def _inicializar_excel(self):
        """Crear archivo Excel de seguimiento si no existe"""
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Seguimiento_Constancias"
            
            # Encabezados extendidos con validación completa
            headers = ["ID", "Alumno", "Codigo", "DNI", "Correo", "Carrera", "Ciclo", 
                      "Documento", "Estado", "Autoridad", "Firma", "Fecha_Creacion",
                      "Validado_UNI", "Validado_DNI", "Fuente_Datos", "Facultad", "Estado_UNI"]
            ws.append(headers)
            
            wb.save(self.excel_file)
            print("✅ Archivo de seguimiento Excel creado")
    
    def generar_constancia_completa(self, nombre, codigo, carrera, ciclo, dni=None, correo=None):
        """Flujo RPA completo para generar constancia con validación DNI + UNI"""
        print(f"🔄 Iniciando flujo RPA para {nombre}...")
        
        # 1. VALIDAR DNI (si se proporciona)
        validacion_dni = None
        if dni:
            print("🆔 Validando DNI...")
            validacion_dni = self._validar_dni(dni)
            if not validacion_dni.get('success'):
                print(f"⚠️ DNI no validado: {validacion_dni.get('error')}")
                # Continuar sin DNI (opcional)
        
        # 2. VALIDAR ESTUDIANTE EN PORTAL UNI
        print("🎓 Validando estudiante en portal UNI...")
        validacion_uni = self._validar_estudiante_uni(codigo, nombre)
        
        if not validacion_uni.get('success'):
            print(f"❌ Estudiante no validado en UNI: {validacion_uni.get('error')}")
            return {
                'success': False,
                'error': f"Estudiante no encontrado en portal UNI: {validacion_uni.get('error')}",
                'validacion_uni': validacion_uni,
                'validacion_dni': validacion_dni
            }
        
        print(f"✅ Estudiante validado en UNI: {validacion_uni.get('nombre')}")
        
        # 3. Validar correo institucional
        if correo and not self._validar_correo_uni(correo):
            print(f"⚠️ Correo no es institucional: {correo}")
        
        # Usar datos validados (más confiables)
        datos_validados = self._combinar_datos_validados_completo(
            nombre, codigo, carrera, ciclo, dni, correo, validacion_uni, validacion_dni
        )
        
        # 4. Ejecutar navegador demo (opcional)
        self._ejecutar_navegador_demo()
        
        # 5. Generar PDF con datos validados
        archivo_pdf = self._generar_pdf_constancia_completo(datos_validados)
        
        # 6. Registrar en Excel con información completa de validación
        registro_id = self._registrar_en_excel_completo(datos_validados, archivo_pdf, validacion_uni, validacion_dni)
        
        print("✅ Flujo RPA completado exitosamente con validación completa")
        
        return {
            'success': True,
            'archivo_pdf': archivo_pdf,
            'registro_id': registro_id,
            'datos_validados': datos_validados,
            'validacion_uni': validacion_uni,
            'validacion_dni': validacion_dni
        }
    
    def _ejecutar_navegador_demo(self):
        """Demostración de automatización con navegador usando Selenium"""
        browser = None
        try:
            print("🌐 Abriendo navegador para demostración...")
            
            # Configurar ChromeDriver
            service = Service(ChromeDriverManager().install())
            browser = webdriver.Chrome(service=service)
            
            # Abrir Google como demo
            browser.get("https://www.google.com")
            print("✅ Navegador abierto correctamente")
            
            time.sleep(1)
            
            # Buscar el campo de búsqueda y realizar búsqueda
            search_box = browser.find_element(By.NAME, "q")
            search_box.send_keys("constancia académica universidad")
            search_box.send_keys(Keys.RETURN)
            print("✅ Búsqueda de constancia académica realizada")
            
            time.sleep(2)
            
            # Cerrar navegador
            browser.quit()
            print("✅ Navegador cerrado correctamente")
            
        except Exception as e:
            print(f"⚠️ Error en navegador (continuando): {e}")
            if browser:
                try:
                    browser.quit()
                except:
                    pass
    
    def _generar_pdf_constancia(self, nombre, codigo, carrera, ciclo):
        """Generar PDF de constancia académica"""
        # Crear nombre único para el archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_pdf = f"constancia_{codigo}_{timestamp}.pdf"
        ruta_completa = os.path.join(self.pdf_folder, archivo_pdf)
        
        # Crear PDF con reportlab
        c = canvas.Canvas(ruta_completa, pagesize=letter)
        width, height = letter
        
        # Encabezado
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(width/2, height-100, "UNIVERSIDAD NACIONAL DE INGENIERÍA")
        
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(width/2, height-140, "CONSTANCIA DE ESTUDIOS")
        
        # Contenido
        c.setFont("Helvetica", 12)
        y_position = height - 200
        
        contenido = [
            f"Por medio de la presente se hace constar que:",
            "",
            f"Nombre del Estudiante: {nombre}",
            f"Código de Alumno: {codigo}",
            f"Carrera: {carrera}",
            f"Ciclo Académico: {ciclo}",
            "",
            f"Se encuentra matriculado y cursando estudios regulares",
            f"en esta casa de estudios superiores.",
            "",
            f"Se expide la presente constancia a solicitud del interesado",
            f"para los fines que estime conveniente.",
            "",
            f"Lima, {datetime.now().strftime('%d de %B de %Y')}"
        ]
        
        for linea in contenido:
            c.drawString(100, y_position, linea)
            y_position -= 25
        
        # Pie de página
        c.setFont("Helvetica-Oblique", 10)
        c.drawCentredString(width/2, 150, "Firma pendiente de autoridad competente")
        c.drawCentredString(width/2, 130, "Coordinación Académica")
        
        c.save()
        
        print(f"✅ PDF generado: {archivo_pdf}")
        return archivo_pdf
    
    def _registrar_en_excel(self, nombre, codigo, carrera, ciclo, archivo_pdf):
        """Registrar constancia en Excel de seguimiento"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Generar ID único
            registro_id = str(uuid.uuid4())[:8]
            
            # Agregar nueva fila
            nueva_fila = [
                registro_id,
                nombre,
                codigo,
                carrera,
                ciclo,
                archivo_pdf,
                "Enviado",
                "Coordinador Académico",
                "Pendiente",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
            
            ws.append(nueva_fila)
            wb.save(self.excel_file)
            
            print("✅ Constancia enviada a autoridad")
            print("✅ Seguimiento actualizado en Excel")
            
            return registro_id
            
        except Exception as e:
            print(f"❌ Error al registrar en Excel: {e}")
            raise
    
    def _validar_estudiante_uni(self, codigo, nombre):
        """Validar estudiante en portal UNI"""
        try:
            return self.uni_validator.validar_estudiante_uni(codigo, nombre)
        except Exception as e:
            print(f"⚠️ Error validando en UNI: {e}")
            return {
                'success': False,
                'error': str(e),
                'codigo_buscado': codigo
            }
    
    def _combinar_datos_validados(self, nombre_input, codigo_input, carrera_input, ciclo_input, validacion_uni):
        """Combinar datos de entrada con datos validados de UNI"""
        datos_combinados = {
            'nombre': nombre_input,
            'codigo': codigo_input,
            'carrera': carrera_input,
            'ciclo': ciclo_input,
            'fuente': 'input_usuario'
        }
        
        # Si la validación UNI fue exitosa, usar esos datos como prioritarios
        if validacion_uni.get('success'):
            if validacion_uni.get('nombre'):
                datos_combinados['nombre'] = validacion_uni['nombre']
                datos_combinados['fuente'] = 'uni_validado'
            
            if validacion_uni.get('carrera'):
                datos_combinados['carrera'] = validacion_uni['carrera']
            
            if validacion_uni.get('facultad'):
                datos_combinados['facultad'] = validacion_uni['facultad']
            
            datos_combinados['estado_uni'] = validacion_uni.get('estado', 'Activo')
            datos_combinados['validado_uni'] = True
        else:
            datos_combinados['validado_uni'] = False
            datos_combinados['error_validacion'] = validacion_uni.get('error')
        
        return datos_combinados
    
    def _registrar_en_excel_validado(self, datos_validados, archivo_pdf, validacion_uni):
        """Registrar en Excel con información de validación UNI"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Generar ID único
            registro_id = str(uuid.uuid4())[:8]
            
            # Agregar nueva fila con información extendida
            nueva_fila = [
                registro_id,
                datos_validados['nombre'],
                datos_validados['codigo'],
                datos_validados['carrera'],
                datos_validados['ciclo'],
                archivo_pdf,
                "Enviado",
                "Coordinador Académico",
                "Pendiente",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                # Nuevas columnas para validación UNI
                "SÍ" if datos_validados.get('validado_uni') else "NO",
                datos_validados.get('fuente', 'input_usuario'),
                datos_validados.get('facultad', ''),
                datos_validados.get('estado_uni', ''),
                validacion_uni.get('coincidencia_nombres', 100) if validacion_uni.get('success') else 0
            ]
            
            ws.append(nueva_fila)
            wb.save(self.excel_file)
            
            print("✅ Constancia enviada a autoridad")
            print("✅ Seguimiento actualizado en Excel con validación UNI")
            
            return registro_id
            
        except Exception as e:
            print(f"❌ Error al registrar en Excel validado: {e}")
            raise
    
    def _validar_dni(self, dni):
        """Validar DNI"""
        try:
            return self.dni_validator.validar_dni(dni)
        except Exception as e:
            print(f"⚠️ Error validando DNI: {e}")
            return {
                'success': False,
                'error': str(e),
                'dni_buscado': dni
            }
    
    def _validar_correo_uni(self, correo):
        """Validar que el correo sea institucional @uni.pe"""
        if not correo:
            return False
        return correo.lower().endswith('@uni.pe')
    
    def _combinar_datos_validados_completo(self, nombre, codigo, carrera, ciclo, dni, correo, validacion_uni, validacion_dni):
        """Combinar todos los datos validados"""
        datos_combinados = {
            'nombre': nombre,
            'codigo': codigo,
            'dni': dni or '',
            'correo': correo or '',
            'carrera': carrera,
            'ciclo': ciclo,
            'fuente': 'input_usuario'
        }
        
        # Priorizar datos de UNI si están disponibles
        if validacion_uni and validacion_uni.get('success'):
            if validacion_uni.get('nombre'):
                datos_combinados['nombre'] = validacion_uni['nombre']
                datos_combinados['fuente'] = 'uni_validado'
            
            if validacion_uni.get('carrera'):
                datos_combinados['carrera'] = validacion_uni['carrera']
            
            datos_combinados['validado_uni'] = True
            datos_combinados['estado_uni'] = validacion_uni.get('estado', 'Activo')
            datos_combinados['facultad'] = validacion_uni.get('facultad', '')
        else:
            datos_combinados['validado_uni'] = False
        
        # Usar datos de DNI si están disponibles
        if validacion_dni and validacion_dni.get('success'):
            if validacion_dni.get('nombre_completo'):
                # Solo usar DNI si no hay datos de UNI
                if not datos_combinados.get('validado_uni'):
                    datos_combinados['nombre'] = validacion_dni['nombre_completo']
                    datos_combinados['fuente'] = 'dni_validado'
            
            datos_combinados['validado_dni'] = True
            datos_combinados['nombre_dni'] = validacion_dni.get('nombre_completo', '')
        else:
            datos_combinados['validado_dni'] = False
        
        return datos_combinados
    
    def _generar_pdf_constancia_completo(self, datos_validados):
        """Generar PDF con datos completos validados"""
        # Usar el método existente pero con datos mejorados
        return self._generar_pdf_constancia(
            datos_validados['nombre'],
            datos_validados['codigo'],
            datos_validados['carrera'],
            datos_validados['ciclo']
        )
    
    def _registrar_en_excel_completo(self, datos_validados, archivo_pdf, validacion_uni, validacion_dni):
        """Registrar en Excel con información completa"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Generar ID único
            registro_id = str(uuid.uuid4())[:8]
            
            # Agregar nueva fila con información completa (ORDEN CORRECTO)
            nueva_fila = [
                registro_id,                                    # 0: ID
                datos_validados['nombre'],                      # 1: Alumno
                datos_validados['codigo'],                      # 2: Código
                datos_validados.get('dni', ''),               # 3: DNI
                datos_validados.get('correo', ''),            # 4: Correo
                datos_validados['carrera'],                    # 5: Carrera
                datos_validados['ciclo'],                      # 6: Ciclo
                archivo_pdf,                                   # 7: Documento (ARCHIVO PDF)
                "Enviado",                                     # 8: Estado
                "Coordinador Académico",                       # 9: Autoridad
                "Pendiente",                                   # 10: Firma
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),  # 11: Fecha
                # Nuevas columnas de validación
                "SÍ" if datos_validados.get('validado_uni') else "NO",  # 12: UNI Validado
                "SÍ" if datos_validados.get('validado_dni') else "NO",  # 13: DNI Validado
                datos_validados.get('fuente', 'input_usuario'),         # 14: Fuente
                datos_validados.get('facultad', 'FIEE'),                # 15: Facultad
                datos_validados.get('estado_uni', 'Activo')             # 16: Estado UNI
            ]
            
            ws.append(nueva_fila)
            wb.save(self.excel_file)
            
            print("✅ Constancia enviada a autoridad")
            print("✅ Seguimiento actualizado en Excel con validación completa")
            
            return registro_id
            
        except Exception as e:
            print(f"❌ Error al registrar en Excel completo: {e}")
            raise
    
    def eliminar_constancia(self, registro_id):
        """Eliminar constancia del seguimiento y archivo PDF"""
        try:
            print(f"🗑️ Eliminando constancia: {registro_id}")
            
            if not os.path.exists(self.excel_file):
                print(f"❌ Archivo Excel no encontrado: {self.excel_file}")
                return False
            
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Buscar y eliminar registro
            fila_a_eliminar = None
            archivo_pdf = None
            alumno_nombre = None
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == registro_id:
                    fila_a_eliminar = row_num
                    alumno_nombre = row[1].value if len(row) > 1 else "Desconocido"
                    archivo_pdf = row[7].value if len(row) > 7 else None  # Columna Documento
                    print(f"📋 Encontrado registro: {alumno_nombre} - {archivo_pdf}")
                    break
            
            if fila_a_eliminar:
                # Eliminar fila del Excel
                ws.delete_rows(fila_a_eliminar)
                wb.save(self.excel_file)
                print(f"✅ Registro eliminado del Excel: fila {fila_a_eliminar}")
                
                # Eliminar archivo PDF en múltiples ubicaciones
                if archivo_pdf:
                    possible_paths = [
                        os.path.join(self.pdf_folder, archivo_pdf),
                        os.path.join('autoridad_entrada', archivo_pdf),
                        os.path.join('PDFs', archivo_pdf),
                        os.path.join('constancias', archivo_pdf)
                    ]
                    
                    pdf_eliminado = False
                    for pdf_path in possible_paths:
                        if os.path.exists(pdf_path):
                            try:
                                os.remove(pdf_path)
                                print(f"✅ Archivo PDF eliminado: {pdf_path}")
                                pdf_eliminado = True
                                break
                            except Exception as e:
                                print(f"⚠️ Error eliminando {pdf_path}: {e}")
                    
                    if not pdf_eliminado:
                        print(f"⚠️ Archivo PDF no encontrado para eliminar: {archivo_pdf}")
                
                print(f"✅ Constancia eliminada completamente: {registro_id}")
                return True
            else:
                print(f"❌ Constancia no encontrada: {registro_id}")
                return False
                
        except Exception as e:
            print(f"❌ Error eliminando constancia: {e}")
            raise
    
    def obtener_seguimiento(self):
        """Obtener lista de constancias para mostrar en web"""
        try:
            if not os.path.exists(self.excel_file):
                return []
            
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            constancias = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay ID
                    constancias.append({
                        'id': row[0],           # 0: ID
                        'alumno': row[1],       # 1: Alumno
                        'codigo': row[2],       # 2: Código
                        'dni': row[3] if len(row) > 3 else '',        # 3: DNI
                        'correo': row[4] if len(row) > 4 else '',     # 4: Correo
                        'carrera': row[5] if len(row) > 5 else '',    # 5: Carrera
                        'ciclo': row[6] if len(row) > 6 else '',      # 6: Ciclo
                        'documento': row[7] if len(row) > 7 else '',  # 7: Documento (PDF)
                        'estado': row[8] if len(row) > 8 else '',     # 8: Estado
                        'autoridad': row[9] if len(row) > 9 else '',  # 9: Autoridad
                        'firma': row[10] if len(row) > 10 else '',    # 10: Firma
                        'fecha': row[11] if len(row) > 11 else ''     # 11: Fecha
                    })
            
            return constancias
            
        except Exception as e:
            print(f"❌ Error al obtener seguimiento: {e}")
            return []
    
    def firmar_constancia(self, registro_id):
        """Simular firma digital de autoridad"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Buscar y actualizar registro
            for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == registro_id:
                    ws.cell(row=row_num, column=9).value = "Firmado"  # Columna Firma
                    ws.cell(row=row_num, column=7).value = "Firmado y Aprobado"  # Columna Estado
                    break
            
            wb.save(self.excel_file)
            print(f"✅ Constancia {registro_id} firmada digitalmente")
            
            return True
            
        except Exception as e:
            print(f"❌ Error al firmar constancia: {e}")
            raise